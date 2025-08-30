#!/usr/bin/env python3
"""
Simple TikTok Search Tool
A minimal tool to search TikTok videos and save links to Excel
"""

import openpyxl
import time
import sys
from urllib.parse import quote
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

class TikTokSearch:
    """Simple TikTok search tool that saves results to Excel"""
    
    def __init__(self):
        """Initialize the search tool"""
        self.driver = None
        self.setup_driver()
    
    def setup_driver(self):
        """Setup Chrome driver with proper options"""
        try:
            chrome_options = Options()
            chrome_options.add_argument("--headless")  # Run in background
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
            
            # Install and setup Chrome driver
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
        except Exception as e:
            print(f"❌ Error setting up Chrome driver: {e}")
            print("💡 Make sure Chrome browser is installed on your system")
            raise
    
    def search_tiktok(self, query, max_results=20):
        """
        Search for TikTok videos using Selenium to handle dynamic content
        
        Args:
            query (str): Search term
            max_results (int): Maximum number of results to find
            
        Returns:
            list: List of dictionaries with video info
        """
        print(f"🔍 Searching TikTok for: {query}")
        
        videos = []
        
        try:
            # Construct search URL
            search_url = f"https://www.tiktok.com/search?q={quote(query)}"
            
            print(f"📡 Loading: {search_url}")
            
            # Navigate to the search page
            self.driver.get(search_url)
            
            # Wait for the page to load (wait for content to appear)
            print("⏳ Waiting for page to load...")
            wait = WebDriverWait(self.driver, 30)
            
            # Wait for video content to load
            try:
                # Wait for any video elements to appear
                wait.until(EC.presence_of_element_located((By.TAG_NAME, "a")))
                print("✅ Page loaded successfully")
            except:
                print("⚠️  Page load timeout, but continuing...")
            
            # Give extra time for dynamic content to load
            print("⏳ Waiting for dynamic content...")
            time.sleep(10)
            
            # Scroll down to load more content
            print("📜 Scrolling to load more content...")
            for i in range(3):
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(3)
            
            # Get the page source after everything is loaded
            page_source = self.driver.page_source
            
            # Look for video links in the fully loaded page
            print("🔍 Extracting video links...")
            
            # Multiple patterns to find TikTok video URLs
            patterns = [
                r'https://www\.tiktok\.com/@[\w.-]+/video/\d+',
                r'https://vm\.tiktok\.com/[A-Za-z0-9]+/',
                r'https://www\.tiktok\.com/t/[A-Za-z0-9]+/',
                r'"url":"(https://www\.tiktok\.com/@[\w.-]+/video/\d+)"',
                r'"shareUrl":"(https://www\.tiktok\.com/@[\w.-]+/video/\d+)"',
                r'href="(https://www\.tiktok\.com/@[\w.-]+/video/\d+)"',
            ]
            
            found_links = []
            for pattern in patterns:
                matches = re.findall(pattern, page_source)
                found_links.extend(matches)
            
            # Remove duplicates while preserving order
            unique_links = []
            seen = set()
            for link in found_links:
                if link not in seen:
                    unique_links.append(link)
                    seen.add(link)
            
            print(f"✅ Found {len(unique_links)} unique video links")
            
            if unique_links:
                # Limit results
                unique_links = unique_links[:max_results]
                
                # Extract additional info for each video
                for i, link in enumerate(unique_links, 1):
                    print(f"📹 Processing video {i}/{len(unique_links)}")
                    
                    # Extract username and video ID from URL
                    username = "unknown"
                    video_id = "unknown"
                    
                    # Pattern 1: Full TikTok URL
                    match = re.search(r'https://www\.tiktok\.com/@([\w.-]+)/video/(\d+)', link)
                    if match:
                        username = match.group(1)
                        video_id = match.group(2)
                    
                    # Pattern 2: Short TikTok URL
                    elif 'vm.tiktok.com' in link or 'tiktok.com/t/' in link:
                        # For short URLs, we'll use the full URL as ID
                        video_id = link.split('/')[-1].replace('/', '')
                        username = "short_url"
                    
                    video_info = {
                        'url': link,
                        'username': username,
                        'video_id': video_id,
                        'title': f"Video by @{username}",
                        'search_query': query
                    }
                    videos.append(video_info)
                
                print(f"🎉 Successfully processed {len(videos)} videos")
            else:
                print("❌ No videos found")
                print("💡 This might be due to TikTok's anti-bot measures")
                print("💡 Try using a different search term or try again later")
            
            return videos
            
        except Exception as e:
            print(f"❌ Error during search: {e}")
            return []
    
    def save_to_excel(self, videos, filename="tiktok_search_results.xlsx"):
        """
        Save video links to Excel file
        
        Args:
            videos (list): List of video dictionaries
            filename (str): Excel filename
        """
        if not videos:
            print("❌ No videos to save")
            return
        
        print(f"💾 Saving {len(videos)} videos to {filename}")
        
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TikTok Videos"
        
        # Add headers
        headers = ['URL', 'Username', 'Video ID', 'Title', 'Search Query']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row, video in enumerate(videos, 2):
            ws.cell(row=row, column=1, value=video['url'])
            ws.cell(row=row, column=2, value=video['username'])
            ws.cell(row=row, column=3, value=video['video_id'])
            ws.cell(row=row, column=4, value=video['title'])
            ws.cell(row=row, column=5, value=video['search_query'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the file
        try:
            wb.save(filename)
            print(f"✅ Successfully saved to {filename}")
            print(f"📊 File contains {len(videos)} video links")
        except Exception as e:
            print(f"❌ Error saving Excel file: {e}")
    
    def search_and_save(self, query, max_results=20, filename=None):
        """
        Search for videos and save to Excel
        
        Args:
            query (str): Search term
            max_results (int): Maximum number of results
            filename (str): Excel filename (optional)
        """
        print(f"🚀 Starting TikTok search for: {query}")
        
        # Generate filename if not provided
        if not filename:
            safe_query = re.sub(r'[^\w\s-]', '', query).strip()
            safe_query = re.sub(r'[-\s]+', '_', safe_query)
            filename = f"tiktok_search_{safe_query}.xlsx"
        
        # Search for videos
        videos = self.search_tiktok(query, max_results)
        
        if videos:
            # Save to Excel
            self.save_to_excel(videos, filename)
            print(f"\n🎯 Search complete! Found {len(videos)} videos")
            print(f"📁 Results saved to: {filename}")
        else:
            print("❌ No videos found")
    
    def cleanup(self):
        """Clean up resources"""
        if self.driver:
            self.driver.quit()

def main():
    """Main function to run the TikTok search tool"""
    print("🎵 Simple TikTok Search Tool")
    print("=" * 50)
    print("🔍 This tool searches TikTok and saves video links to Excel")
    print("=" * 50)
    
    # Get search query from user
    if len(sys.argv) > 1:
        query = " ".join(sys.argv[1:])
    else:
        query = input("Enter search term: ").strip()
    
    if not query:
        print("❌ No search term provided")
        print("\n💡 Usage examples:")
        print("   python tiktok_search.py 'dance videos'")
        print("   python tiktok_search.py 'funny cats'")
        print("   python tiktok_search.py 'cooking tutorial'")
        return
    
    # Get number of results
    try:
        max_results = int(input("Maximum number of results (default 20): ") or "20")
    except ValueError:
        max_results = 20
    
    # Create search tool and run
    searcher = TikTokSearch()
    try:
        searcher.search_and_save(query, max_results)
    finally:
        searcher.cleanup()

if __name__ == "__main__":
    main()
