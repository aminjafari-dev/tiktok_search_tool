"""
Excel management for TikTok Search Tool
Handles Excel file creation, formatting, and data saving
"""

import os
import openpyxl
from src.core.config import EXCEL_CONFIG, MESSAGES


class ExcelManager:
    """Manages Excel file operations for TikTok search results"""
    
    def __init__(self):
        """Initialize the Excel manager"""
        self.workbook = None
        self.worksheet = None
        self.existing_links = set()  # Track existing links to avoid duplicates
    
    def load_existing_workbook(self, filename):
        """
        Load an existing Excel workbook if it exists
        
        Args:
            filename (str): Path to the Excel file
            
        Returns:
            bool: True if file exists and was loaded, False otherwise
        """
        if os.path.exists(filename):
            try:
                self.workbook = openpyxl.load_workbook(filename)
                self.worksheet = self.workbook.active
                
                # Extract existing links to avoid duplicates
                self._extract_existing_links()
                
                print(f"📂 Loaded existing file: {filename}")
                print(f"📊 Found {len(self.existing_links)} existing links")
                return True
            except Exception as e:
                print(f"⚠️  Error loading existing file: {e}")
                print("🔄 Creating new file instead")
                return False
        return False
    
    def _extract_existing_links(self):
        """
        Extract existing video URLs from the worksheet to track duplicates
        
        This method reads all existing video URLs from the first column
        and stores them in a set for fast duplicate checking.
        """
        if not self.worksheet:
            return
        
        self.existing_links.clear()
        
        # Start from row 2 (after headers) and read all URLs from column 1
        for row in range(2, self.worksheet.max_row + 1):
            cell_value = self.worksheet.cell(row=row, column=1).value
            if cell_value and isinstance(cell_value, str):
                # Clean the URL and add to set
                clean_url = cell_value.strip()
                if clean_url:
                    self.existing_links.add(clean_url)
    
    def filter_duplicate_videos(self, videos):
        """
        Filter out videos that already exist in the Excel file
        
        Args:
            videos (list): List of video dictionaries to filter
            
        Returns:
            tuple: (new_videos, duplicate_count) - filtered videos and count of duplicates
        """
        new_videos = []
        duplicate_count = 0
        
        for video in videos:
            video_url = video.get('url', '').strip()
            if video_url and video_url not in self.existing_links:
                new_videos.append(video)
                # Add to existing links set to prevent duplicates within the same batch
                self.existing_links.add(video_url)
            else:
                duplicate_count += 1
        
        return new_videos, duplicate_count
    
    def create_workbook(self, sheet_name=None):
        """
        Create a new Excel workbook and worksheet
        
        Args:
            sheet_name (str): Name for the worksheet (optional)
        """
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        
        # Set worksheet name
        if sheet_name:
            self.worksheet.title = sheet_name
        else:
            self.worksheet.title = EXCEL_CONFIG["sheet_name"]
    
    def add_headers(self, headers=None):
        """
        Add headers to the worksheet (only if they don't exist)
        
        Args:
            headers (list): List of header strings (optional)
        """
        if headers is None:
            headers = EXCEL_CONFIG["headers"]
        
        # Only add headers if the worksheet is empty (row 1 is empty)
        if self.worksheet.max_row == 0 or not self.worksheet.cell(row=1, column=1).value:
            for col, header in enumerate(headers, 1):
                self.worksheet.cell(row=1, column=col, value=header)
    
    def add_data(self, data, start_row=None):
        """
        Add data rows to the worksheet
        
        Args:
            data (list): List of dictionaries containing video data
            start_row (int): Starting row number (optional, auto-calculated if None)
        """
        if start_row is None:
            # Start after the last existing row
            start_row = self.worksheet.max_row + 1
        
        for row, video in enumerate(data, start_row):
            # Map video data to columns
            self.worksheet.cell(row=row, column=1, value=video.get('url', ''))
            self.worksheet.cell(row=row, column=2, value=video.get('username', ''))
            self.worksheet.cell(row=row, column=3, value=video.get('video_id', ''))
            self.worksheet.cell(row=row, column=4, value=video.get('title', ''))
            self.worksheet.cell(row=row, column=5, value=video.get('search_query', ''))
            self.worksheet.cell(row=row, column=6, value=video.get('added_date', ''))
    
    def auto_adjust_columns(self, max_width=None):
        """
        Automatically adjust column widths based on content
        
        Args:
            max_width (int): Maximum column width (optional)
        """
        if max_width is None:
            max_width = EXCEL_CONFIG["max_column_width"]
        
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, max_width)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def save_workbook(self, filename):
        """
        Save the workbook to a file
        
        Args:
            filename (str): Output filename
            
        Returns:
            bool: True if save successful, False otherwise
        """
        try:
            # Ensure the directory exists
            directory = os.path.dirname(filename)
            if directory and not os.path.exists(directory):
                os.makedirs(directory)
                print(f"📁 Created directory: {directory}")
            
            self.workbook.save(filename)
            return True
        except Exception as e:
            print(f"❌ Error saving Excel file: {e}")
            return False
    
    def create_and_save(self, videos, filename, headers=None):
        """
        Create workbook, add data, and save in one operation
        
        Args:
            videos (list): List of video dictionaries
            filename (str): Output filename
            headers (list): Custom headers (optional)
            
        Returns:
            bool: True if successful, False otherwise
        """
        if not videos:
            print("❌ No videos to save")
            return False
        
        print(MESSAGES["saving"].format(count=len(videos), filename=filename))
        
        try:
            # Try to load existing workbook first
            file_exists = self.load_existing_workbook(filename)
            
            if not file_exists:
                # Create new workbook and worksheet
                self.create_workbook()
            
            # Add headers (only if new file)
            if not file_exists:
                self.add_headers(headers)
            
            # Filter out duplicate videos
            new_videos, duplicate_count = self.filter_duplicate_videos(videos)
            
            if duplicate_count > 0:
                print(f"⚠️  Skipped {duplicate_count} duplicate links")
            
            if new_videos:
                # Add new data
                self.add_data(new_videos)
                
                # Auto-adjust columns
                self.auto_adjust_columns()
                
                # Save file
                if self.save_workbook(filename):
                    print(MESSAGES["saved"].format(filename=filename))
                    print(f"📊 Added {len(new_videos)} new links to existing file")
                    print(f"📈 Total links in file: {len(self.existing_links)}")
                    return True
                else:
                    return False
            else:
                print("ℹ️  All videos already exist in the file - no new data added")
                return True
                
        except Exception as e:
            print(f"❌ Error creating Excel file: {e}")
            return False
    
    def cleanup(self):
        """Clean up workbook resources"""
        if self.workbook:
            try:
                self.workbook.close()
            except:
                pass
