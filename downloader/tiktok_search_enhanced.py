"""
TikTok Video Search and Excel Export Module (Enhanced Version)

This module provides functionality to search TikTok videos by subject/keyword
and save the video links to a single Excel file with duplicate prevention.

Features:
- Single Excel file that gets updated each time
- Duplicate link prevention
- Date tracking for when links were added
- Configurable link count (50-60 per run)
- Automatic scrolling to load more content

Usage:
    # Search for videos and add to existing Excel file
    python tiktok_search_enhanced.py --subject "baby fun moments" --min-links 50 --max-links 60
    
    # Programmatic usage
    from tiktok_search_enhanced import TikTokSearcherEnhanced
    searcher = TikTokSearcherEnhanced()
    searcher.search_and_update_excel("baby fun moments", min_links=50, max_links=60)
"""

import os
import sys
import time
import logging
import argparse
from pathlib import Path
from typing import List, Dict, Optional, Any, Set
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from colorama import init, Fore, Style

# Initialize colorama for cross-platform colored output
init(autoreset=True)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('tiktok_search_enhanced.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class TikTokSearcherEnhanced:
    """
    Enhanced TikTok searcher that maintains a single Excel file with unique links.
    
    This class provides methods to search TikTok for videos based on keywords,
    extract video information and links, and save the results to a single Excel file
    with duplicate prevention and date tracking.
    
    Attributes:
        output_dir (str): Directory where Excel files will be saved
        excel_filename (str): Name of the Excel file to maintain
        min_links (int): Minimum number of links to find per search
        max_links (int): Maximum number of links to find per search
        headless (bool): Whether to run browser in headless mode
        delay (float): Delay between actions to avoid rate limiting
    """
    
    def __init__(self, output_dir: str = "downloads", excel_filename: str = "tiktok_videos_master.xlsx",
                 min_links: int = 50, max_links: int = 60, headless: bool = True, delay: float = 2.0):
        """
        Initialize the enhanced TikTok searcher with specified options.
        
        Args:
            output_dir (str): Directory to save Excel files (default: "downloads")
            excel_filename (str): Name of the Excel file to maintain (default: "tiktok_videos_master.xlsx")
            min_links (int): Minimum links to find per search (default: 50)
            max_links (int): Maximum links to find per search (default: 60)
            headless (bool): Run browser in headless mode (default: True)
            delay (float): Delay between actions in seconds (default: 2.0)
        """
        self.output_dir = Path(output_dir)
        self.excel_filename = excel_filename
        self.min_links = min_links
        self.max_links = max_links
        self.headless = headless
        self.delay = delay
        
        # Create output directory if it doesn't exist
        self.output_dir.mkdir(exist_ok=True)
        
        # Excel file path
        self.excel_filepath = self.output_dir / self.excel_filename
        
        # Set to track existing video IDs to avoid duplicates
        self.existing_video_ids: Set[str] = set()
        
        # Initialize web driver
        self.driver = None
        self._setup_driver()
        
        # Load existing video IDs from Excel file
        self._load_existing_video_ids()
    
    def _setup_driver(self):
        """
        Set up the Chrome web driver with appropriate options.
        
        This method configures the Chrome driver with user agent spoofing,
        headless mode if enabled, and other options to avoid detection.
        """
        try:
            # Configure Chrome options
            chrome_options = Options()
            
            if self.headless:
                chrome_options.add_argument("--headless")
            
            # Add user agent to avoid detection (using a standard one)
            chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            
            # Additional options to avoid detection
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            # Set up Chrome service
            service = Service(ChromeDriverManager().install())
            
            # Initialize driver
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
            # Execute script to remove webdriver property
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            logger.info("Chrome driver initialized successfully")
            print(f"{Fore.GREEN}Chrome driver initialized successfully{Style.RESET_ALL}")
            
        except Exception as e:
            logger.error(f"Failed to initialize Chrome driver: {e}")
            print(f"{Fore.RED}Error: Failed to initialize Chrome driver: {e}{Style.RESET_ALL}")
            raise
    
    def _load_existing_video_ids(self):
        """
        Load existing video IDs from the Excel file to prevent duplicates.
        """
        if self.excel_filepath.exists():
            try:
                wb = load_workbook(self.excel_filepath)
                ws = wb.active
                
                # Read existing video IDs from the first column
                for row in range(2, ws.max_row + 1):  # Skip header row
                    video_id = ws.cell(row=row, column=1).value
                    if video_id:
                        self.existing_video_ids.add(str(video_id))
                
                print(f"{Fore.CYAN}Loaded {len(self.existing_video_ids)} existing video IDs{Style.RESET_ALL}")
                logger.info(f"Loaded {len(self.existing_video_ids)} existing video IDs")
                
            except Exception as e:
                logger.warning(f"Could not load existing video IDs: {e}")
                print(f"{Fore.YELLOW}Warning: Could not load existing video IDs{Style.RESET_ALL}")
        else:
            print(f"{Fore.CYAN}No existing Excel file found. Will create new one.{Style.RESET_ALL}")
    
    def search_videos(self, subject: str) -> List[Dict[str, Any]]:
        """
        Search TikTok for videos based on the given subject.
        
        Args:
            subject (str): The subject/keyword to search for
            
        Returns:
            List[Dict[str, Any]]: List of video information dictionaries
        """
        if not self.driver:
            logger.error("Web driver not initialized")
            return []
        
        videos = []
        search_url = f"https://www.tiktok.com/search?q={subject.replace(' ', '%20')}"
        
        try:
            print(f"{Fore.CYAN}Searching TikTok for: '{subject}'{Style.RESET_ALL}")
            logger.info(f"Searching TikTok for subject: {subject}")
            
            # Navigate to search page
            self.driver.get(search_url)
            time.sleep(self.delay)
            
            # Wait for page to load
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            
            # Scroll to load more videos until we have enough unique ones
            videos = self._scroll_and_extract_videos(subject)
            
            print(f"{Fore.GREEN}Found {len(videos)} new unique videos for subject: '{subject}'{Style.RESET_ALL}")
            logger.info(f"Found {len(videos)} new unique videos for subject: {subject}")
            
        except Exception as e:
            logger.error(f"Error during search: {e}")
            print(f"{Fore.RED}Error during search: {e}{Style.RESET_ALL}")
        
        return videos
    
    def _scroll_and_extract_videos(self, subject: str) -> List[Dict[str, Any]]:
        """
        Scroll down the page and extract videos until we have enough unique ones.
        
        Args:
            subject (str): The search subject
            
        Returns:
            List[Dict[str, Any]]: List of unique video information dictionaries
        """
        videos = []
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        scroll_attempts = 0
        max_scroll_attempts = 20  # Increased for more content
        
        print(f"{Fore.YELLOW}Loading videos by scrolling (target: {self.min_links}-{self.max_links} unique videos)...{Style.RESET_ALL}")
        
        while len(videos) < self.max_links and scroll_attempts < max_scroll_attempts:
            # Extract videos from current page
            new_videos = self._extract_video_links_from_current_page(subject)
            
            # Add only unique videos
            for video in new_videos:
                video_id = video.get('video_id')
                if video_id and video_id not in self.existing_video_ids and len(videos) < self.max_links:
                    videos.append(video)
                    self.existing_video_ids.add(video_id)
            
            print(f"{Fore.CYAN}Found {len(videos)} unique videos so far...{Style.RESET_ALL}")
            
            # Scroll down
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(self.delay)
            
            # Calculate new scroll height
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            
            # If height is the same, we've reached the bottom
            if new_height == last_height:
                scroll_attempts += 1
            else:
                scroll_attempts = 0
                last_height = new_height
            
            print(f"{Fore.CYAN}Scroll attempt {scroll_attempts + 1}/{max_scroll_attempts}{Style.RESET_ALL}")
            
            # If we have enough videos, we can stop
            if len(videos) >= self.min_links:
                print(f"{Fore.GREEN}Reached minimum target of {self.min_links} videos{Style.RESET_ALL}")
                break
        
        return videos
    
    def _extract_video_links_from_current_page(self, subject: str) -> List[Dict[str, Any]]:
        """
        Extract video links and information from the current page.
        
        Args:
            subject (str): The search subject
            
        Returns:
            List[Dict[str, Any]]: List of video information dictionaries
        """
        videos = []
        
        try:
            # Look for video links (TikTok video URLs)
            video_elements = self.driver.find_elements(By.CSS_SELECTOR, "a[href*='/video/']")
            
            for element in video_elements:
                try:
                    href = element.get_attribute("href")
                    if href and "/video/" in href:
                        # Extract video information
                        video_info = self._extract_video_info(element, href, subject)
                        if video_info:
                            videos.append(video_info)
                
                except Exception as e:
                    logger.warning(f"Error extracting video info: {e}")
                    continue
            
        except Exception as e:
            logger.error(f"Error extracting video links: {e}")
        
        return videos
    
    def _extract_video_info(self, element, href: str, subject: str) -> Optional[Dict[str, Any]]:
        """
        Extract detailed information about a video from its element.
        
        Args:
            element: Selenium web element containing video information
            href (str): Video URL
            subject (str): Search subject
            
        Returns:
            Optional[Dict[str, Any]]: Video information dictionary or None
        """
        try:
            # Extract video ID from URL
            video_id = href.split('/video/')[-1].split('?')[0]
            
            # Skip if we already have this video
            if video_id in self.existing_video_ids:
                return None
            
            # Try to extract title/description
            title = "Unknown"
            try:
                # Look for title in various possible locations
                title_element = element.find_element(By.CSS_SELECTOR, "[data-e2e='search-card-desc']")
                title = title_element.text.strip()
            except NoSuchElementException:
                try:
                    title_element = element.find_element(By.CSS_SELECTOR, "[data-e2e='search-card-title']")
                    title = title_element.text.strip()
                except NoSuchElementException:
                    pass
            
            # Try to extract username
            username = "Unknown"
            try:
                username_element = element.find_element(By.CSS_SELECTOR, "[data-e2e='search-card-user-link']")
                username = username_element.text.strip()
            except NoSuchElementException:
                pass
            
            return {
                'video_id': video_id,
                'url': href,
                'title': title,
                'username': username,
                'subject': subject,
                'search_timestamp': datetime.now().isoformat(),
                'date_added': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
        except Exception as e:
            logger.warning(f"Error extracting video info: {e}")
            return None
    
    def update_excel_file(self, videos: List[Dict[str, Any]]) -> bool:
        """
        Update the Excel file with new videos, maintaining existing data.
        
        Args:
            videos (List[Dict[str, Any]]): List of video information dictionaries
            
        Returns:
            bool: True if successful, False otherwise
        """
        if not videos:
            print(f"{Fore.YELLOW}No new videos to add to Excel{Style.RESET_ALL}")
            return False
        
        try:
            # Load existing workbook or create new one
            if self.excel_filepath.exists():
                wb = load_workbook(self.excel_filepath)
                ws = wb.active
                print(f"{Fore.CYAN}Updating existing Excel file{Style.RESET_ALL}")
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "TikTok Videos"
                
                # Write headers for new file
                headers = ['Video ID', 'URL', 'Title', 'Username', 'Subject', 'Search Timestamp', 'Date Added']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                print(f"{Fore.CYAN}Creating new Excel file{Style.RESET_ALL}")
            
            # Find the next row to write to
            next_row = ws.max_row + 1 if ws.max_row > 1 else 2
            
            # Write new videos
            for video in videos:
                ws.cell(row=next_row, column=1, value=video.get('video_id', ''))
                ws.cell(row=next_row, column=2, value=video.get('url', ''))
                ws.cell(row=next_row, column=3, value=video.get('title', ''))
                ws.cell(row=next_row, column=4, value=video.get('username', ''))
                ws.cell(row=next_row, column=5, value=video.get('subject', ''))
                ws.cell(row=next_row, column=6, value=video.get('search_timestamp', ''))
                ws.cell(row=next_row, column=7, value=video.get('date_added', ''))
                next_row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(self.excel_filepath)
            
            print(f"{Fore.GREEN}Excel file updated: {self.excel_filepath}{Style.RESET_ALL}")
            print(f"{Fore.GREEN}Added {len(videos)} new videos{Style.RESET_ALL}")
            logger.info(f"Excel file updated with {len(videos)} new videos")
            
            return True
            
        except Exception as e:
            logger.error(f"Error updating Excel file: {e}")
            print(f"{Fore.RED}Error updating Excel file: {e}{Style.RESET_ALL}")
            return False
    
    def search_and_update_excel(self, subject: str) -> bool:
        """
        Search for videos and update Excel file in one operation.
        
        Args:
            subject (str): The subject/keyword to search for
            
        Returns:
            bool: True if successful, False otherwise
        """
        # Search for videos
        videos = self.search_videos(subject)
        
        if videos:
            # Update Excel file
            return self.update_excel_file(videos)
        else:
            print(f"{Fore.YELLOW}No new videos found for subject: '{subject}'{Style.RESET_ALL}")
            return False
    
    def get_excel_stats(self) -> Dict[str, Any]:
        """
        Get statistics about the Excel file.
        
        Returns:
            Dict[str, Any]: Statistics about the Excel file
        """
        if not self.excel_filepath.exists():
            return {"total_videos": 0, "subjects": [], "file_exists": False}
        
        try:
            wb = load_workbook(self.excel_filepath)
            ws = wb.active
            
            total_videos = ws.max_row - 1  # Subtract header row
            subjects = set()
            
            # Collect unique subjects
            for row in range(2, ws.max_row + 1):
                subject = ws.cell(row=row, column=5).value
                if subject:
                    subjects.add(subject)
            
            return {
                "total_videos": total_videos,
                "subjects": list(subjects),
                "file_exists": True,
                "file_path": str(self.excel_filepath)
            }
            
        except Exception as e:
            logger.error(f"Error getting Excel stats: {e}")
            return {"total_videos": 0, "subjects": [], "file_exists": False, "error": str(e)}
    
    def close(self):
        """
        Close the web driver and clean up resources.
        """
        if self.driver:
            self.driver.quit()
            logger.info("Web driver closed")
            print(f"{Fore.YELLOW}Web driver closed{Style.RESET_ALL}")


def main():
    """
    Main function to handle command-line interface.
    
    This function parses command-line arguments and initiates the search process.
    It supports searching for videos by subject and updating the Excel file.
    """
    parser = argparse.ArgumentParser(
        description="Search TikTok videos by subject and update Excel file with unique links",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python tiktok_search_enhanced.py --subject "baby fun moments" --min-links 50 --max-links 60
  python tiktok_search_enhanced.py --subject "cooking recipes" --min-links 30 --max-links 40
  python tiktok_search_enhanced.py --subject "dance moves" --excel-file "my_videos.xlsx"
        """
    )
    
    # Required arguments
    parser.add_argument(
        '--subject', 
        type=str, 
        required=True,
        help='Subject/keyword to search for on TikTok'
    )
    
    # Optional arguments
    parser.add_argument(
        '--min-links', 
        type=int, 
        default=50,
        help='Minimum number of links to find (default: 50)'
    )
    parser.add_argument(
        '--max-links', 
        type=int, 
        default=60,
        help='Maximum number of links to find (default: 60)'
    )
    parser.add_argument(
        '--output-dir', 
        type=str, 
        default='downloads',
        help='Output directory for Excel files (default: downloads)'
    )
    parser.add_argument(
        '--excel-file', 
        type=str, 
        default='tiktok_videos_master.xlsx',
        help='Excel filename to maintain (default: tiktok_videos_master.xlsx)'
    )
    parser.add_argument(
        '--no-headless', 
        action='store_true',
        help='Run browser in visible mode (not headless)'
    )
    parser.add_argument(
        '--delay', 
        type=float, 
        default=2.0,
        help='Delay between actions in seconds (default: 2.0)'
    )
    parser.add_argument(
        '--stats', 
        action='store_true',
        help='Show Excel file statistics before searching'
    )
    
    args = parser.parse_args()
    
    # Initialize searcher
    searcher = TikTokSearcherEnhanced(
        output_dir=args.output_dir,
        excel_filename=args.excel_file,
        min_links=args.min_links,
        max_links=args.max_links,
        headless=not args.no_headless,
        delay=args.delay
    )
    
    try:
        # Show stats if requested
        if args.stats:
            stats = searcher.get_excel_stats()
            print(f"\n{Fore.CYAN}📊 Excel File Statistics:{Style.RESET_ALL}")
            if stats["file_exists"]:
                print(f"📁 File: {stats['file_path']}")
                print(f"📊 Total Videos: {stats['total_videos']}")
                print(f"🏷️  Subjects: {', '.join(stats['subjects'])}")
            else:
                print("📁 No Excel file found yet")
            print()
        
        # Search and update Excel
        success = searcher.search_and_update_excel(args.subject)
        
        if success:
            print(f"{Fore.GREEN}✅ Successfully updated Excel file with new video links!{Style.RESET_ALL}")
            
            # Show final stats
            final_stats = searcher.get_excel_stats()
            print(f"{Fore.CYAN}📊 Final Statistics:{Style.RESET_ALL}")
            print(f"📊 Total Videos: {final_stats['total_videos']}")
            print(f"🏷️  Subjects: {', '.join(final_stats['subjects'])}")
        else:
            print(f"{Fore.YELLOW}⚠️  No new videos found or failed to update Excel file{Style.RESET_ALL}")
    
    except KeyboardInterrupt:
        print(f"\n{Fore.YELLOW}Search interrupted by user{Style.RESET_ALL}")
        sys.exit(1)
    except Exception as e:
        print(f"{Fore.RED}Unexpected error: {e}{Style.RESET_ALL}")
        sys.exit(1)
    finally:
        searcher.close()


if __name__ == "__main__":
    main()
