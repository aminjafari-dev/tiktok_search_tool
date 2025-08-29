"""
TikTok Video Search and Excel Export Module (Basic Version)

This module provides functionality to search TikTok videos by subject/keyword
and save the video links to Excel files. Basic version without external dependencies.

Usage:
    # Search for videos by subject and save to Excel
    python tiktok_search_basic.py --subject "cooking recipes" --max-videos 50
    
    # Programmatic usage
    from tiktok_search_basic import TikTokSearcher
    searcher = TikTokSearcher()
    links = searcher.search_videos("cooking recipes", max_videos=50)
    searcher.save_to_excel(links, "cooking_videos.xlsx")
"""

import os
import sys
import time
import logging
import argparse
from pathlib import Path
from typing import List, Dict, Optional, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from colorama import init, Fore, Style

# Initialize colorama for cross-platform colored output
init(autoreset=True)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('tiktok_search.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class TikTokSearcher:
    """
    A class to search TikTok videos by subject and extract video links to Excel.
    
    This class provides methods to search TikTok for videos based on keywords,
    extract video information and links, and save the results to Excel files.
    No video downloading is performed - only link extraction and Excel export.
    
    Attributes:
        output_dir (str): Directory where Excel files will be saved
        max_videos (int): Maximum number of videos to search for
        headless (bool): Whether to run browser in headless mode
        delay (float): Delay between actions to avoid rate limiting
    """
    
    def __init__(self, output_dir: str = "downloads", max_videos: int = 50, 
                 headless: bool = True, delay: float = 2.0):
        """
        Initialize the TikTok searcher with specified options.
        
        Args:
            output_dir (str): Directory to save Excel files (default: "downloads")
            max_videos (int): Maximum videos to search for (default: 50)
            headless (bool): Run browser in headless mode (default: True)
            delay (float): Delay between actions in seconds (default: 2.0)
        """
        self.output_dir = Path(output_dir)
        self.max_videos = max_videos
        self.headless = headless
        self.delay = delay
        
        # Create output directory if it doesn't exist
        self.output_dir.mkdir(exist_ok=True)
        
        # Initialize web driver
        self.driver = None
        self._setup_driver()
    
    def _setup_driver(self):
        """
        Set up the Chrome web driver with appropriate options.
        
        This method configures the Chrome driver with user agent spoofing,
        headless mode if enabled, and other options to avoid detection.
        """
        try:
            # Configure Chrome options
            chrome_options = Options()
            
            if self.headless:
                chrome_options.add_argument("--headless")
            
            # Add user agent to avoid detection (using a standard one)
            chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            
            # Additional options to avoid detection
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            # Set up Chrome service
            service = Service(ChromeDriverManager().install())
            
            # Initialize driver
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            
            # Execute script to remove webdriver property
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            logger.info("Chrome driver initialized successfully")
            print(f"{Fore.GREEN}Chrome driver initialized successfully{Style.RESET_ALL}")
            
        except Exception as e:
            logger.error(f"Failed to initialize Chrome driver: {e}")
            print(f"{Fore.RED}Error: Failed to initialize Chrome driver: {e}{Style.RESET_ALL}")
            raise
    
    def search_videos(self, subject: str) -> List[Dict[str, Any]]:
        """
        Search TikTok for videos based on the given subject.
        
        Args:
            subject (str): The subject/keyword to search for
            
        Returns:
            List[Dict[str, Any]]: List of video information dictionaries
        """
        if not self.driver:
            logger.error("Web driver not initialized")
            return []
        
        videos = []
        search_url = f"https://www.tiktok.com/search?q={subject.replace(' ', '%20')}"
        
        try:
            print(f"{Fore.CYAN}Searching TikTok for: '{subject}'{Style.RESET_ALL}")
            logger.info(f"Searching TikTok for subject: {subject}")
            
            # Navigate to search page
            self.driver.get(search_url)
            time.sleep(self.delay)
            
            # Wait for page to load
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            
            # Scroll to load more videos
            self._scroll_to_load_videos()
            
            # Extract video links
            videos = self._extract_video_links()
            
            print(f"{Fore.GREEN}Found {len(videos)} videos for subject: '{subject}'{Style.RESET_ALL}")
            logger.info(f"Found {len(videos)} videos for subject: {subject}")
            
        except Exception as e:
            logger.error(f"Error during search: {e}")
            print(f"{Fore.RED}Error during search: {e}{Style.RESET_ALL}")
        
        return videos
    
    def _scroll_to_load_videos(self):
        """
        Scroll down the page to load more videos dynamically.
        
        This method scrolls the page multiple times to ensure all available
        videos are loaded before extraction.
        """
        print(f"{Fore.YELLOW}Loading videos by scrolling...{Style.RESET_ALL}")
        
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        scroll_attempts = 0
        max_attempts = 10
        
        while scroll_attempts < max_attempts:
            # Scroll down
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(self.delay)
            
            # Calculate new scroll height
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            
            # If height is the same, we've reached the bottom
            if new_height == last_height:
                scroll_attempts += 1
            else:
                scroll_attempts = 0
                last_height = new_height
            
            print(f"{Fore.CYAN}Scroll attempt {scroll_attempts + 1}/{max_attempts}{Style.RESET_ALL}")
    
    def _extract_video_links(self) -> List[Dict[str, Any]]:
        """
        Extract video links and information from the search results page.
        
        Returns:
            List[Dict[str, Any]]: List of video information dictionaries
        """
        videos = []
        
        try:
            # Look for video links (TikTok video URLs)
            video_elements = self.driver.find_elements(By.CSS_SELECTOR, "a[href*='/video/']")
            
            for element in video_elements[:self.max_videos]:
                try:
                    href = element.get_attribute("href")
                    if href and "/video/" in href:
                        # Extract video information
                        video_info = self._extract_video_info(element, href)
                        if video_info:
                            videos.append(video_info)
                            
                            if len(videos) >= self.max_videos:
                                break
                
                except Exception as e:
                    logger.warning(f"Error extracting video info: {e}")
                    continue
            
            # Remove duplicates based on video ID
            unique_videos = []
            seen_ids = set()
            
            for video in videos:
                video_id = video.get('video_id')
                if video_id and video_id not in seen_ids:
                    unique_videos.append(video)
                    seen_ids.add(video_id)
            
            return unique_videos
            
        except Exception as e:
            logger.error(f"Error extracting video links: {e}")
            return videos
    
    def _extract_video_info(self, element, href: str) -> Optional[Dict[str, Any]]:
        """
        Extract detailed information about a video from its element.
        
        Args:
            element: Selenium web element containing video information
            href (str): Video URL
            
        Returns:
            Optional[Dict[str, Any]]: Video information dictionary or None
        """
        try:
            # Extract video ID from URL
            video_id = href.split('/video/')[-1].split('?')[0]
            
            # Try to extract title/description
            title = "Unknown"
            try:
                # Look for title in various possible locations
                title_element = element.find_element(By.CSS_SELECTOR, "[data-e2e='search-card-desc']")
                title = title_element.text.strip()
            except NoSuchElementException:
                try:
                    title_element = element.find_element(By.CSS_SELECTOR, "[data-e2e='search-card-title']")
                    title = title_element.text.strip()
                except NoSuchElementException:
                    pass
            
            # Try to extract username
            username = "Unknown"
            try:
                username_element = element.find_element(By.CSS_SELECTOR, "[data-e2e='search-card-user-link']")
                username = username_element.text.strip()
            except NoSuchElementException:
                pass
            
            return {
                'video_id': video_id,
                'url': href,
                'title': title,
                'username': username,
                'search_timestamp': datetime.now().isoformat(),
                'subject': getattr(self, '_current_subject', 'Unknown')
            }
            
        except Exception as e:
            logger.warning(f"Error extracting video info: {e}")
            return None
    
    def save_to_excel(self, videos: List[Dict[str, Any]], filename: str = None) -> str:
        """
        Save video information to an Excel file using openpyxl.
        
        Args:
            videos (List[Dict[str, Any]]): List of video information dictionaries
            filename (str): Output filename (optional, will auto-generate if None)
            
        Returns:
            str: Path to the created Excel file
        """
        if not videos:
            print(f"{Fore.YELLOW}No videos to save to Excel{Style.RESET_ALL}")
            return ""
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            subject = getattr(self, '_current_subject', 'tiktok').replace(' ', '_')
            filename = f"{subject}_videos_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "TikTok Videos"
            
            # Define headers
            headers = ['Video ID', 'URL', 'Title', 'Username', 'Subject', 'Search Timestamp']
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row, video in enumerate(videos, 2):
                ws.cell(row=row, column=1, value=video.get('video_id', ''))
                ws.cell(row=row, column=2, value=video.get('url', ''))
                ws.cell(row=row, column=3, value=video.get('title', ''))
                ws.cell(row=row, column=4, value=video.get('username', ''))
                ws.cell(row=row, column=5, value=video.get('subject', ''))
                ws.cell(row=row, column=6, value=video.get('search_timestamp', ''))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filepath)
            
            print(f"{Fore.GREEN}Excel file saved: {filepath}{Style.RESET_ALL}")
            logger.info(f"Excel file saved: {filepath}")
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error saving to Excel: {e}")
            print(f"{Fore.RED}Error saving to Excel: {e}{Style.RESET_ALL}")
            return ""
    
    def search_and_save(self, subject: str, filename: str = None) -> str:
        """
        Search for videos and save to Excel in one operation.
        
        Args:
            subject (str): The subject/keyword to search for
            filename (str): Output filename (optional)
            
        Returns:
            str: Path to the created Excel file
        """
        # Store current subject for filename generation
        self._current_subject = subject
        
        # Search for videos
        videos = self.search_videos(subject)
        
        if videos:
            # Save to Excel
            return self.save_to_excel(videos, filename)
        else:
            print(f"{Fore.YELLOW}No videos found for subject: '{subject}'{Style.RESET_ALL}")
            return ""
    
    def close(self):
        """
        Close the web driver and clean up resources.
        """
        if self.driver:
            self.driver.quit()
            logger.info("Web driver closed")
            print(f"{Fore.YELLOW}Web driver closed{Style.RESET_ALL}")


def main():
    """
    Main function to handle command-line interface.
    
    This function parses command-line arguments and initiates the search process.
    It supports searching for videos by subject and saving results to Excel.
    """
    parser = argparse.ArgumentParser(
        description="Search TikTok videos by subject and save links to Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python tiktok_search_basic.py --subject "cooking recipes" --max-videos 50
  python tiktok_search_basic.py --subject "dance" --max-videos 100 --no-headless
  python tiktok_search_basic.py --subject "comedy" --output-file "comedy_videos.xlsx"
        """
    )
    
    # Required arguments
    parser.add_argument(
        '--subject', 
        type=str, 
        required=True,
        help='Subject/keyword to search for on TikTok'
    )
    
    # Optional arguments
    parser.add_argument(
        '--max-videos', 
        type=int, 
        default=50,
        help='Maximum number of videos to search for (default: 50)'
    )
    parser.add_argument(
        '--output-dir', 
        type=str, 
        default='downloads',
        help='Output directory for Excel files (default: downloads)'
    )
    parser.add_argument(
        '--output-file', 
        type=str, 
        help='Specific output filename (optional, auto-generated if not provided)'
    )
    parser.add_argument(
        '--no-headless', 
        action='store_true',
        help='Run browser in visible mode (not headless)'
    )
    parser.add_argument(
        '--delay', 
        type=float, 
        default=2.0,
        help='Delay between actions in seconds (default: 2.0)'
    )
    
    args = parser.parse_args()
    
    # Initialize searcher
    searcher = TikTokSearcher(
        output_dir=args.output_dir,
        max_videos=args.max_videos,
        headless=not args.no_headless,
        delay=args.delay
    )
    
    try:
        # Search and save to Excel
        excel_file = searcher.search_and_save(args.subject, args.output_file)
        
        if excel_file:
            print(f"{Fore.GREEN}✅ Successfully created Excel file with video links!{Style.RESET_ALL}")
            print(f"{Fore.CYAN}📁 File location: {excel_file}{Style.RESET_ALL}")
        else:
            print(f"{Fore.YELLOW}⚠️  No videos found or failed to save Excel file{Style.RESET_ALL}")
    
    except KeyboardInterrupt:
        print(f"\n{Fore.YELLOW}Search interrupted by user{Style.RESET_ALL}")
        sys.exit(1)
    except Exception as e:
        print(f"{Fore.RED}Unexpected error: {e}{Style.RESET_ALL}")
        sys.exit(1)
    finally:
        searcher.close()


if __name__ == "__main__":
    main()

