"""
Excel management for TikTok Search Tool
Handles Excel file creation, formatting, and data saving
"""

import openpyxl
from config import EXCEL_CONFIG, MESSAGES


class ExcelManager:
    """Manages Excel file operations for TikTok search results"""
    
    def __init__(self):
        """Initialize the Excel manager"""
        self.workbook = None
        self.worksheet = None
    
    def create_workbook(self, sheet_name=None):
        """
        Create a new Excel workbook and worksheet
        
        Args:
            sheet_name (str): Name for the worksheet (optional)
        """
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        
        # Set worksheet name
        if sheet_name:
            self.worksheet.title = sheet_name
        else:
            self.worksheet.title = EXCEL_CONFIG["sheet_name"]
    
    def add_headers(self, headers=None):
        """
        Add headers to the worksheet
        
        Args:
            headers (list): List of header strings (optional)
        """
        if headers is None:
            headers = EXCEL_CONFIG["headers"]
        
        for col, header in enumerate(headers, 1):
            self.worksheet.cell(row=1, column=col, value=header)
    
    def add_data(self, data, start_row=2):
        """
        Add data rows to the worksheet
        
        Args:
            data (list): List of dictionaries containing video data
            start_row (int): Starting row number (default: 2, after headers)
        """
        for row, video in enumerate(data, start_row):
            # Map video data to columns
            self.worksheet.cell(row=row, column=1, value=video.get('url', ''))
            self.worksheet.cell(row=row, column=2, value=video.get('username', ''))
            self.worksheet.cell(row=row, column=3, value=video.get('video_id', ''))
            self.worksheet.cell(row=row, column=4, value=video.get('title', ''))
            self.worksheet.cell(row=row, column=5, value=video.get('search_query', ''))
    
    def auto_adjust_columns(self, max_width=None):
        """
        Automatically adjust column widths based on content
        
        Args:
            max_width (int): Maximum column width (optional)
        """
        if max_width is None:
            max_width = EXCEL_CONFIG["max_column_width"]
        
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, max_width)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def save_workbook(self, filename):
        """
        Save the workbook to a file
        
        Args:
            filename (str): Output filename
            
        Returns:
            bool: True if save successful, False otherwise
        """
        try:
            self.workbook.save(filename)
            return True
        except Exception as e:
            print(f"❌ Error saving Excel file: {e}")
            return False
    
    def create_and_save(self, videos, filename, headers=None):
        """
        Create workbook, add data, and save in one operation
        
        Args:
            videos (list): List of video dictionaries
            filename (str): Output filename
            headers (list): Custom headers (optional)
            
        Returns:
            bool: True if successful, False otherwise
        """
        if not videos:
            print("❌ No videos to save")
            return False
        
        print(MESSAGES["saving"].format(count=len(videos), filename=filename))
        
        try:
            # Create workbook and worksheet
            self.create_workbook()
            
            # Add headers
            self.add_headers(headers)
            
            # Add data
            self.add_data(videos)
            
            # Auto-adjust columns
            self.auto_adjust_columns()
            
            # Save file
            if self.save_workbook(filename):
                print(MESSAGES["saved"].format(filename=filename))
                print(MESSAGES["file_info"].format(count=len(videos)))
                return True
            else:
                return False
                
        except Exception as e:
            print(f"❌ Error creating Excel file: {e}")
            return False
    
    def cleanup(self):
        """Clean up workbook resources"""
        if self.workbook:
            try:
                self.workbook.close()
            except:
                pass
